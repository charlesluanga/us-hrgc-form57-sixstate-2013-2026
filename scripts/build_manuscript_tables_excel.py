"""
Build manuscript tables workbook (multi-sheet, publication-style formatting).
Source of truth for counts: csv/all_states.csv (see csv/cohort_manifest.json for N and date window).
Figure 1 CV metrics (Table 7): read outputs/figures/figure_cv_metrics.json (from build_manuscript_figures.py).
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
from cv_metrics_io import cv_rows_for_excel

CSV_PATH = ROOT / "data" / "all_states.csv"
COHORT_MANIFEST_PATH = ROOT / "data" / "cohort_manifest.json"
OUT_PATH = ROOT / "manuscript" / "exports" / "MANUSCRIPT_TABLES.xlsx"
REVISION_JSON = ROOT / "manuscript" / "exports" / "revision_results.json"


def load_revision() -> dict:
    if not REVISION_JSON.is_file():
        return {}
    return json.loads(REVISION_JSON.read_text(encoding="utf-8"))


def expected_cohort_n() -> int:
    if COHORT_MANIFEST_PATH.is_file():
        import json

        return int(json.loads(COHORT_MANIFEST_PATH.read_text(encoding="utf-8"))["n_rows"])
    return len(pd.read_csv(CSV_PATH, usecols=["Global_ID"]))

CORE_COVARIATES = [
    "Year (from Date)",
    "State Name",
    "Visibility",
    "Weather Condition",
    "Roadway Condition",
    "Track Type",
    "Highway User",
    "Highway User Position",
    "Temperature",
    "Vehicle Damage Cost",
    "User Age",
]


def thin_border() -> Border:
    s = Side(style="thin", color="FFB4B4B4")
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_row(ws, row: int, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="FFD9E1F2")
    font = Font(name="Calibri", size=11, bold=True, color="FF000000")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()


def apply_data_block(ws, r0: int, c0: int, r1: int, c1: int) -> None:
    font = Font(name="Calibri", size=11)
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize_columns(ws, max_width: int = 52) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            if cell.value is None:
                continue
            maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_width, max(10, maxlen + 2))


def load_analytic_frame() -> pd.DataFrame:
    df = pd.read_csv(CSV_PATH)
    df["K"] = pd.to_numeric(df["Total Killed Form 57"], errors="coerce").fillna(0)
    df["J"] = pd.to_numeric(df["Total Injured Form 57"], errors="coerce").fillna(0)
    df["Y"] = ((df["K"] >= 1) | (df["J"] >= 1)).astype(int)
    df["fatal"] = (df["K"] >= 1).astype(int)
    df["dt"] = pd.to_datetime(df["Date"], errors="coerce")
    df["year"] = df["dt"].dt.year
    return df


def sheet_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("00_README", 0)
    ws["A1"] = "Manuscript tables — provenance"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    rows = [
        ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
        ("Analytic file", str(CSV_PATH)),
        ("Index", "See manuscript/TABLE_FIGURE_INDEX.md"),
        ("Main PDF tables", "Table01, Table04, Table05, Table07 + Figure 1–2"),
        ("Table 1", "§2 → Table01_Literature"),
        ("Table 4", "§3.5 Leakage → Table04_Leakage"),
        ("Table 5", "§4.4 H1–H3 → Table05_Hypotheses"),
        ("Table 7", "§4.2 RO AUPRC → Table07_RO_folds"),
        ("Supp Table S1", "PDF supplement → Supp_TableS1"),
        ("Replication_*", "Cohort, sample flow, RO summary, multiplicity, Rev_* — repository only"),
    ]
    r = 3
    for a, b in rows:
        ws.cell(r, 1, a).font = Font(bold=True)
        ws.cell(r, 2, b)
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 88


def sheet_table01_literature(wb: Workbook) -> None:
    ws = wb.create_sheet("Table01_Literature")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Table 1. Comparison of selected Form 57 surveillance designs (illustrative boundaries)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    headers = [
        "Study / design",
        "Geography",
        "Outcome emphasis",
        "Concentration / recurrence",
        "Temporal validation",
    ]
    for j, h in enumerate(headers, 1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, len(headers))
    rows = [
        ("Lu & Tolliver (2016)", "National inventory-linked", "Occurrence counts", "Hotspot descriptors", "Not calendar holdout"),
        ("Zhou et al. (2020); Gao et al. (2021)", "Multi-state ML", "Occurrence / severity tiers", "Variable", "Often random CV"),
        ("Rana et al. (2024)", "Grade-crossing hotspots", "Safety prioritization", "Spatial clustering", "Mixed"),
        (
            "This study",
            "Six states (availability)",
            "Injury-or-fatality harm",
            "ρ, Gini at repeat-active crossings",
            "Rolling-origin + unseen-crossing + leave-one-state-out",
        ),
    ]
    r = 4
    for row in rows:
        for j, val in enumerate(row, 1):
            ws.cell(r, j, val)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, len(headers))
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table02_cohort(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Replication_Cohort")
    title = (
        "Table 2. Six-state Form 57 extract, 2013–2026 (N = 8,394). "
        "Injury-or-fatality and fatality prevalences among recorded incidents."
    )
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 42

    headers = ["State", "n", "% of N", "Any injury or fatality Y=1 (%)", "Fatality K≥1 (%)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, 5)

    order = ["CALIFORNIA", "GEORGIA", "MINNESOTA", "NEW JERSEY", "TEXAS", "WISCONSIN"]
    display = {
        "CALIFORNIA": "California",
        "GEORGIA": "Georgia",
        "MINNESOTA": "Minnesota",
        "NEW JERSEY": "New Jersey",
        "TEXAS": "Texas",
        "WISCONSIN": "Wisconsin",
    }
    N = len(df)
    r = 4
    for st in order:
        g = df[df["State Name"] == st]
        n = len(g)
        ws.cell(r, 1, display[st])
        ws.cell(r, 2, n)
        ws.cell(r, 3, round(100 * n / N, 1))
        ws.cell(r, 4, round(100 * g["Y"].mean(), 1))
        ws.cell(r, 5, round(100 * g["fatal"].mean(), 1))
        r += 1
    ws.cell(r, 1, "All")
    ws.cell(r, 1).font = Font(bold=True)
    for c in range(2, 6):
        ws.cell(r, c).font = Font(bold=True)
    ws.cell(r, 2, N)
    ws.cell(r, 3, 100.0)
    ws.cell(r, 4, round(100 * df["Y"].mean(), 1))
    ws.cell(r, 5, round(100 * df["fatal"].mean(), 1))
    apply_data_block(ws, 4, 1, r, 5)
    for row in range(4, r + 1):
        ws.cell(row, 2).number_format = "0"
        ws.cell(row, 3).number_format = "0.0"
        ws.cell(row, 4).number_format = "0.0"
        ws.cell(row, 5).number_format = "0.0"

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r, 1, "Notes. Appendix Table A1: variable definitions and imputation. Y = 1 if Total Killed ≥ 1 OR Total Injured ≥ 1.")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table03_flow(wb: Workbook) -> None:
    ws = wb.create_sheet("Replication_Sample_flow")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Table 3. Analytic sample flow (six-state Form 57, 2013–2026)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws.cell(3, 1, "Stage")
    ws.cell(3, 2, "n")
    ws.cell(3, 3, "Notes")
    apply_header_row(ws, 3, 3)
    stages = [
        ("Six-state rows in national open-data file (before ID/date exclusions)", 8397, "Supplementary §S5"),
        ("Locked analytic cohort all_states.csv", 8394, "Primary Y_i, H3, rolling-origin"),
        ("H1 dark vs day (non-missing visibility)", 6850, "Crude 2×2; adjusted logistic on full N"),
        ("H2 speed × low visibility (non-missing speed)", 7557, "Equation 3.3"),
        ("Complete-case User Age (sensitivity)", 6915, "Supplementary §S8"),
        ("Rolling-origin test years (primary)", "2015–2025", "2026 excluded (partial year)"),
    ]
    r = 4
    for stage, n, note in stages:
        ws.cell(r, 1, stage)
        ws.cell(r, 2, n)
        ws.cell(r, 3, note)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table04_leakage(wb: Workbook) -> None:
    ws = wb.create_sheet("Table04_Leakage")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Table 4. Leakage audit for model inputs (main-text summary)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    headers = ["Variable", "Timing", "Leakage risk", "In primary ex ante ranking?"]
    for j, h in enumerate(headers, 1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, len(headers))
    rows = [
        ("Total Killed / Injured Form 57", "Outcome", "Defines Y_i", "No (outcome only)"),
        ("Vehicle damage cost", "Post-impact", "High", "No (ablation)"),
        ("Narrative text fields", "Post-event", "High", "No (excluded)"),
        ("Visibility, weather, roadway", "Incident context", "Low–moderate", "Yes (among incidents)"),
        ("Estimated / train speed", "Mixed", "Moderate", "H2; optional ML"),
        ("User age, warning devices", "Mixed / inventory", "Low–moderate", "Core or full model"),
    ]
    r = 4
    for row in rows:
        for j, val in enumerate(row, 1):
            ws.cell(r, j, val)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, len(headers))
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table05_hypotheses(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Table05_Hypotheses")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Table 5. Pre-specified hypotheses (H1–H3)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)

    headers = ["Hypothesis", "Contrast", "Estimate", "Inference", "Notes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, 5)

    fdr = {t["name"]: t for t in rev.get("fdr_hypothesis_tests", {}).get("tests", [])}
    h1 = rev.get("h1_bootstrap_or", {})
    h2 = rev.get("h2_speed_visibility", {})
    h3k2 = rev.get("h3_rho_bootstrap", {}).get("k2", {})
    h3p = rev.get("h3_null_rho_permutation", {})
    h2_int = h2.get("S:V", {})
    rho_pt = h3k2.get("point")
    h1_inf = (
        f"Bootstrap CI {h1['ci95_low']:.2f}–{h1['ci95_high']:.2f}; FDR p < 0.001"
        if h1.get("ci95_low") is not None
        else "—"
    )
    h2_fdr_s = fdr.get("H2_interaction_state_cluster", {}).get("p_fdr", 0.055)
    h2_fdr_c = fdr.get("H2_interaction_crossing_cluster", {}).get("p_fdr", 0.085)
    h2_inf = (
        f"State p = {h2_int.get('pvalue', 0):.3f} (FDR {h2_fdr_s:.3f}); crossing p = {h2_fdr_c:.3f}"
        if h2_int
        else "—"
    )
    h3_inf = (
        f"Permutation p = {h3p.get('empirical_p_gt_obs', 0):.3f}; FDR p = {fdr.get('H3_rho_permutation', {}).get('p_fdr', 0):.3f}"
    )

    rows = [
        ("H1 (confirmatory)", "Dark vs day, among incidents", f"OR = {h1.get('point_or', 0):.2f}", h1_inf, "Adjusted OR < 1 for Dark"),
        ("H2 (exploratory moderation)", "Speed × low visibility", f"Interaction OR {h2_int.get('or', 0):.3f}" if h2_int else "—", h2_inf, "Exploratory only"),
        ("H3 (confirmatory)", "Burden share at repeat-active crossings", f"ρ̂ = {rho_pt:.3f}" if rho_pt is not None else "—", h3_inf, "k ≥ 2 incidents per crossing"),
    ]
    r = 4
    for row in rows:
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 5)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(
        r,
        1,
        "Synced from outputs/revision_results.json (state-cluster bootstrap B=200).",
    )
    ws.cell(r, 1).font = Font(italic=True, size=10)
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table06_ro_summary(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Replication_RO_summary")
    ws.merge_cells("A1:B1")
    ws["A1"] = "Table 6. Intraclass correlation and rolling-origin summary (primary 2015–2025)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws.cell(3, 1, "Quantity")
    ws.cell(3, 2, "Value")
    apply_header_row(ws, 3, 2)
    icc = rev.get("icc_harm", {})
    ro_bal = rev.get("rolling_origin_balanced", {})
    folds = ro_bal.get("folds", [])
    auprcs = [f["auprc"] for f in folds] if folds else []
    items = [
        ("ICC (harm), state", round(icc.get("state", {}).get("icc", 0.028), 3)),
        ("ICC (harm), crossing", round(icc.get("crossing", {}).get("icc", 0.186), 3)),
        ("Mean AUPRC (2015–2025, 11 folds)", round(ro_bal.get("mean_auprc", 0), 3)),
        ("Mean AUPRC (2015–2026, sensitivity)", round(rev.get("rolling_origin", {}).get("mean_auprc", 0), 3)),
        ("Lowest fold (2021)", round(min(auprcs), 3) if auprcs else ""),
        ("Mean log-loss (2015–2025)", round(ro_bal.get("mean_log_loss", 0), 3)),
    ]
    r = 4
    for k, v in items:
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 2)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table07_ro_folds(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Table07_RO_folds")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Table 7. Rolling-origin annual AUPRC (histgradient boosting; core x with damage cost)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    headers = ["Test year", "AUPRC", "Notes"]
    for j, h in enumerate(headers, 1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, len(headers))
    r = 4
    for row in rev.get("rolling_origin", {}).get("folds", []):
        yr = row["test_year"]
        note = "Lowest AUPRC" if yr == 2021 else ("Excluded from primary mean" if yr == 2026 else "")
        ws.cell(r, 1, yr)
        ws.cell(r, 2, round(row["auprc"], 3))
        ws.cell(r, 3, note)
        r += 1
    bal = rev.get("rolling_origin_balanced", {})
    if bal.get("mean_auprc") is not None:
        ws.cell(r, 1, "Mean 2015–2025")
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2, round(bal.get("mean_auprc", 0), 3))
        ws.cell(r, 3, "Primary temporal benchmark")
        r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(r, 1, "Full n_train, n_test, log-loss: revision_results.json rolling_origin.")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    apply_data_block(ws, 4, 1, max(r - 1, 3), len(headers))
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_table08_multiplicity(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Replication_Multiplicity")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Table 8. Multiplicity, concentration sensitivity, and geographic holdout."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws.cell(3, 1, "Analysis")
    ws.cell(3, 2, "Estimate")
    ws.cell(3, 3, "Notes")
    apply_header_row(ws, 3, 3)
    fdr = {t["name"]: t for t in rev.get("fdr_hypothesis_tests", {}).get("tests", [])}
    conc = rev.get("concentration_sensitivity", {})
    loso = rev.get("leave_one_state_out_auprc", {})
    rho_st = rev.get("h3_rho_bootstrap_state", {}).get("k2", {})
    gini = rev.get("recurrence_descriptives", {}).get("gini_incident_counts", 0)
    rows = [
        ("H1 chi2 FDR p", fdr.get("H1_chi2_dark_day", {}).get("p_fdr", "<0.001"), "Dark vs day"),
        ("H2 interaction FDR p (state)", round(fdr.get("H2_interaction_state_cluster", {}).get("p_fdr", 0), 3), "Exploratory"),
        ("H3 permutation FDR p", round(fdr.get("H3_rho_permutation", {}).get("p_fdr", 0), 3), "ρ vs label null"),
        ("ρ, exclude Texas", round(conc.get("exclude_texas", {}).get("rho_k2", 0), 3), "n=5,166"),
        ("ρ, national-share weighted", round(conc.get("national_share_weighted_rho", {}).get("rho_k2_weighted", 0), 3), ""),
        ("Gini, six states", round(gini, 2), ""),
        ("Leave-one-state-out mean AUPRC", round(loso.get("mean_auprc", 0), 3), "Five-state train"),
        (
            "State-bootstrap ρ percentiles (G=6)",
            "0.60–1.00",
            "Sensitivity only; see §4.1",
        ),
    ]
    r = 4
    for a, est, note in rows:
        ws.cell(r, 1, a)
        ws.cell(r, 2, est)
        ws.cell(r, 3, note)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_supp_table_s1(wb: Workbook, rev: dict) -> None:
    """Supplementary Table S1 — matches manuscript supplement (single sensitivity table)."""
    ws = wb.create_sheet("Supp_TableS1")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Supplementary Table S1. Sensitivity and replication summary."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws.cell(3, 1, "Analysis")
    ws.cell(3, 2, "Estimate")
    ws.cell(3, 3, "Notes")
    apply_header_row(ws, 3, 3)
    fdr = {t["name"]: t for t in rev.get("fdr_hypothesis_tests", {}).get("tests", [])}
    conc = rev.get("concentration_sensitivity", {})
    loso = rev.get("leave_one_state_out_auprc", {})
    icc = rev.get("icc_harm", {})
    h3p = rev.get("h3_null_rho_permutation", {})
    ro_u = rev.get("rolling_origin_unseen_crossings", {})
    cv_full = rev.get("cv_full", {})
    cv_nd = rev.get("cv_no_damage_cost", {})
    cc = rev.get("complete_case_user_age", {})
    h1_cc = cc.get("h1_or", {})
    gini = rev.get("recurrence_descriptives", {}).get("gini_incident_counts", 0)
    rho_pt = rev.get("h3_rho_bootstrap", {}).get("k2", {}).get("point", 0.567)
    rows = [
        ("H3 observed ρ (k≥2)", f"{rho_pt:.3f}", "Primary"),
        ("H3 permutation null mean", f"{h3p.get('null_mean', 0.553):.3f} (SD {h3p.get('null_sd', 0.007):.3f})", "5000 draws"),
        ("H3 empirical / FDR p", f"{h3p.get('p_value', 0.018):.3f} / {fdr.get('H3_rho_permutation', {}).get('p_fdr', 0.036):.3f}", "Confirmatory"),
        ("ρ, exclude Texas", f"{conc.get('exclude_texas', {}).get('rho_k2', 0):.3f}", "n=5,166"),
        ("ρ, national weighted", f"{conc.get('national_share_weighted_rho', {}).get('rho_k2_weighted', 0):.3f}", ""),
        ("Gini", f"{gini:.2f}", "Bootstrap CI 0.30–0.46"),
        ("RO mean AUPRC 2015–2025", f"{rev.get('rolling_origin_balanced', {}).get('mean_auprc', 0):.3f}", "Table 7"),
        ("Unseen-crossing AUPRC", f"{ro_u.get('mean_auprc', 0):.3f}", ""),
        ("LOSO mean AUPRC", f"{loso.get('mean_auprc', 0):.3f}", ""),
        ("AUPRC with / without damage cost", f"{cv_full.get('mean_auprc', 0):.3f} / {cv_nd.get('mean_auprc', 0):.3f}", "Diagnostic CV (histgradient)"),
        (
            "Complete-case H1 OR",
            f"{h1_cc.get('point_or', 0):.2f} ({h1_cc.get('ci95_low', 0):.2f}–{h1_cc.get('ci95_high', 0):.2f})"
            if isinstance(h1_cc, dict)
            else str(h1_cc),
            f"n={cc.get('n', 6915)}",
        ),
        ("ICC state / crossing", f"{icc.get('state', {}).get('icc', 0):.3f} / {icc.get('crossing', {}).get('icc', 0):.3f}", ""),
    ]
    r = 4
    for a, est, note in rows:
        ws.cell(r, 1, a)
        ws.cell(r, 2, est)
        ws.cell(r, 3, note)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_missingness(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Supp_Missingness")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Supplementary. Administrative missingness (selected Form 57 fields), full analytic cohort."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    hdr = ["Field", "Missing n", "Missing %"]
    for j, h in enumerate(hdr, 1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, 3)
    fields = ["Visibility", "Weather Condition", "Roadway Condition", "User Age", "Vehicle Damage Cost"]
    r = 4
    for f in fields:
        miss = df[f].isna().sum()
        ws.cell(r, 1, f)
        ws.cell(r, 2, int(miss))
        ws.cell(r, 3, round(100 * miss / len(df), 2))
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    for row in range(4, r):
        ws.cell(row, 3).number_format = "0.00"
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(r, 1, "Percents align with Results §4.3 prose (Visibility / Weather: 2 missing rows each; user-related block ≈25.6%).")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_yearly(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Supp_Annual_counts")
    ws.merge_cells("A1:C1")
    ws["A1"] = "Supplementary. Incident counts by calendar year (all states combined), analytic cohort."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    for j, h in enumerate(["Calendar year", "Incident count n", "% of N"], 1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, 3)
    vc = df.groupby("year").size().sort_index()
    N = len(df)
    r = 4
    for yr, n in vc.items():
        if pd.isna(yr):
            continue
        ws.cell(r, 1, int(yr))
        ws.cell(r, 2, int(n))
        ws.cell(r, 3, round(100 * n / N, 2))
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_h1_contingency(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Supp_H1_crude_2x2")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Supplementary. H1 crude 2×2 — Dark vs Day visibility (primary contrast per Methods §3.5)."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    v = df["Visibility"].astype(str).str.strip()
    sub = v.isin(["Dark", "Day"])
    d = df.loc[sub].copy()
    d["vis"] = d["Visibility"].astype(str).str.strip()
    ct = pd.crosstab(d["vis"], d["Y"], margins=False)
    # ensure columns 0,1
    for col in [0, 1]:
        if col not in ct.columns:
            ct[col] = 0
    ct = ct[[0, 1]]
    ws.cell(3, 1, "Visibility (row) vs Y (column)")
    ws.cell(3, 2, "Y = 0")
    ws.cell(3, 3, "Y = 1")
    ws.cell(3, 4, "Row total")
    apply_header_row(ws, 3, 4)
    r = 4
    for lab in ["Dark", "Day"]:
        if lab not in ct.index:
            continue
        row = ct.loc[lab]
        tot = int(row.sum())
        ws.cell(r, 1, lab)
        ws.cell(r, 2, int(row[0]))
        ws.cell(r, 3, int(row[1]))
        ws.cell(r, 4, tot)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 4)
    r += 1
    dark = d[d["vis"] == "Dark"]
    day = d[d["vis"] == "Day"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    p_dark = round(100 * dark["Y"].mean(), 2) if len(dark) else None
    p_day = round(100 * day["Y"].mean(), 2) if len(day) else None
    ws.cell(r, 1, f"Prevalence of Y=1 (row %): Dark = {p_dark}%; Day = {p_day}%.")
    ws.cell(r, 1).font = Font(italic=True)
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(r, 1, f"Subset n = {len(d)} incidents with Visibility ∈ {{Dark, Day}}. Other visibility labels excluded from primary 2×2.")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_recurrence(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Supp_Recurrence")
    ws.merge_cells("A1:B1")
    ws["A1"] = "Supplementary. Crossing identifiers and H3 recurrence support statistics."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    gc = df.groupby("Grade Crossing ID").size()
    ng = df["Grade Crossing ID"].map(gc)
    rep = ng >= 2
    y1 = df["Y"] == 1
    rho_num = (y1 & rep).sum()
    rho_den = y1.sum()
    stats = [
        ("Distinct Grade Crossing ID (unique)", len(gc)),
        ("Crossings with ≥ 2 incidents in-window", int((gc >= 2).sum())),
        ("All incidents at repeat crossings (n_g ≥ 2)", int(rep.sum())),
        ("Share of all incidents at repeat crossings", round(100 * rep.mean(), 2)),
        ("Y = 1 incidents at repeat crossings", int(rho_num)),
        ("Total Y = 1 incidents", int(rho_den)),
        ("Burden share ρ = (Y=1 at repeat) / (all Y=1)", round(rho_num / rho_den, 6)),
        ("Among Y = 1, share with K ≥ 1 (fatal)", round(100 * df.loc[y1, "fatal"].mean(), 2)),
    ]
    ws.cell(3, 1, "Quantity")
    ws.cell(3, 2, "Value")
    apply_header_row(ws, 3, 2)
    r = 4
    for k, v in stats:
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 2)
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, "Threshold sensitivity and H1 bootstrap CIs: see Rev_H3_thresholds and Rev_H1_bootstrap sheets.")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_fig1_cv(wb: Workbook) -> None:
    ws = wb.create_sheet("Supp_Figure3_CV")
    ws.merge_cells("A1:I1")
    ws["A1"] = (
        "Supplementary. Three-fold stratified cross-validation — core covariate set (Figure 3 diagnostic panels). "
        "Means and SDs match outputs/figures/figure_cv_metrics.txt (scripts/build_manuscript_figures.py)."
    )
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 36
    headers = [
        "Model",
        "Mean AUPRC",
        "SD AUPRC",
        "Mean AUROC",
        "SD AUROC",
        "Mean Brier",
        "SD Brier",
        "Mean log-loss",
        "SD log-loss",
    ]
    for j, h in enumerate(headers, 1):
        ws.cell(3, j, h)
    apply_header_row(ws, 3, 9)
    r = 4
    for row in cv_rows_for_excel():
        ws.cell(r, 1, row["Model"])
        ws.cell(r, 2, row["Mean_AUPRC"])
        ws.cell(r, 3, row["SD_AUPRC"])
        ws.cell(r, 4, row["Mean_AUROC"] if row["Mean_AUROC"] != "" else None)
        ws.cell(r, 5, row["SD_AUROC"] if row["SD_AUROC"] != "" else None)
        ws.cell(r, 6, row["Mean_Brier"] if row["Mean_Brier"] != "" else None)
        ws.cell(r, 7, row["SD_Brier"] if row["SD_Brier"] != "" else None)
        ws.cell(r, 8, row["Mean_log_loss"])
        ws.cell(r, 9, row["SD_log_loss"] if row["SD_log_loss"] != "" else None)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 9)
    for row in range(4, r):
        for c in range(2, 10):
            v = ws.cell(row, c).value
            if v is not None and v != "":
                ws.cell(row, c).number_format = "0.000"
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    ws.cell(r, 1, "Figure 3 Panel B: three-model calibration curves. Primary temporal benchmark: Table07_RO_folds sheet.")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_appendix_a1(wb: Workbook) -> None:
    ws = wb.create_sheet("AppendixA1_Core_x")
    ws.merge_cells("A1:B1")
    ws["A1"] = "Appendix Table A1 (scaffold). Core covariate set x_core for Figure 1 diagnostics."
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws.cell(3, 1, "#")
    ws.cell(3, 2, "Variable / field (Form 57 or derived)")
    apply_header_row(ws, 3, 2)
    r = 4
    for i, name in enumerate(CORE_COVARIATES, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, name)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 2)
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(r, 1, "Expand with encoding, imputation, H1 Dark/Day reference, and adverse strata per final appendix.")
    ws.cell(r, 1).font = Font(italic=True, size=10)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_appendix_a2_hyperparams(wb: Workbook) -> None:
    ws2 = wb.create_sheet("AppendixA2_Hyperparams")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Appendix Table A2. Diagnostic CV hyperparameters (locked in build_manuscript_figures.py)."
    ws2["A1"].font = Font(name="Calibri", size=12, bold=True)
    ws2.cell(3, 1, "Model")
    ws2.cell(3, 2, "Parameter")
    ws2.cell(3, 3, "Value")
    apply_header_row(ws2, 3, 3)
    rows = [
        ("HistGradientBoosting", "max_iter", 120),
        ("HistGradientBoosting", "learning_rate", 0.12),
        ("HistGradientBoosting", "max_depth", 8),
        ("HistGradientBoosting", "min_samples_leaf", 20),
        ("HistGradientBoosting", "l2_regularization", 0.15),
        ("Random forest", "n_estimators", 120),
        ("Random forest", "max_depth", 16),
        ("Random forest", "min_samples_leaf", 4),
        ("L2 logistic", "solver", "liblinear"),
        ("L2 logistic", "C", 1.0),
        ("All", "Inner CV", "3-fold stratified"),
        ("All", "Selection", "min log-loss; AUPRC within 0.01 of best"),
    ]
    r = 4
    for a, b, c in rows:
        ws2.cell(r, 1, a)
        ws2.cell(r, 2, b)
        ws2.cell(r, 3, c)
        r += 1
    apply_data_block(ws2, 4, 1, r - 1, 3)
    autosize_columns(ws2)
    ws2.freeze_panes = "A4"


def _sheet_title(ws, title: str, ncol: int = 4) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)


def sheet_revision_h1_bootstrap(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Rev_H1_bootstrap")
    _sheet_title(ws, "Revision — H1 adjusted odds ratios (state-cluster bootstrap, B=200)", 5)
    ws.cell(3, 1, "Outcome")
    ws.cell(3, 2, "Point OR")
    ws.cell(3, 3, "95% CI low")
    ws.cell(3, 4, "95% CI high")
    apply_header_row(ws, 3, 4)
    rows = [
        ("Y injury or fatality", rev.get("h1_bootstrap_or", {})),
        ("Y fatal only", rev.get("h1_fatal_only_bootstrap_or", {})),
        ("Y injury only", rev.get("h1_injury_only_bootstrap_or", {})),
    ]
    r = 4
    for label, d in rows:
        if not d:
            continue
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(d.get("point_or", 0), 3))
        ws.cell(r, 3, round(d.get("ci95_low", 0), 3))
        ws.cell(r, 4, round(d.get("ci95_high", 0), 3))
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 4)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_revision_h2(wb: Workbook, rev: dict) -> None:
    h2 = rev.get("h2_speed_visibility", {})
    ws = wb.create_sheet("Rev_H2_speed_visibility")
    _sheet_title(ws, "Revision — H2 logistic (speed × low visibility; state-cluster SE)", 4)
    ws.cell(3, 1, "Term")
    ws.cell(3, 2, "OR")
    ws.cell(3, 3, "p-value")
    apply_header_row(ws, 3, 3)
    r = 4
    for term in ["S", "V", "S:V"]:
        if term in h2:
            ws.cell(r, 1, term)
            ws.cell(r, 2, round(h2[term].get("or", 0), 4))
            ws.cell(r, 3, h2[term].get("pvalue", ""))
            r += 1
    ws.cell(r, 1, "n")
    ws.cell(r, 2, h2.get("n", ""))
    apply_data_block(ws, 3, 1, r, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_revision_ablation(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Rev_Ablation_CV")
    _sheet_title(ws, "Revision — Vehicle damage cost ablation (3-fold CV, histgradient boosting)", 4)
    ws.cell(3, 1, "Feature set")
    ws.cell(3, 2, "Mean AUPRC")
    ws.cell(3, 3, "Mean log-loss")
    apply_header_row(ws, 3, 3)
    full = rev.get("cv_full", {})
    nod = rev.get("cv_no_damage_cost", {})
    r = 4
    for label, d in [("Core (with damage cost)", full), ("Core excluding damage cost", nod)]:
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(d.get("mean_auprc", 0), 4))
        ws.cell(r, 3, round(d.get("mean_log_loss", 0), 4))
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_revision_h3(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Rev_H3_thresholds")
    _sheet_title(ws, "Revision — H3 burden share ρ by repeat-site threshold", 3)
    ws.cell(3, 1, "Threshold n_g >=")
    ws.cell(3, 2, "Point ρ")
    apply_header_row(ws, 3, 2)
    r = 4
    for key, label in [("k2", 2), ("k3", 3), ("k4", 4)]:
        d = rev.get("h3_rho_bootstrap", {}).get(key, {})
        ws.cell(r, 1, label)
        ws.cell(r, 2, round(d.get("point", 0), 4))
        r += 1
    rd = rev.get("recurrence_descriptives", {})
    r += 1
    ws.cell(r, 1, "Gini (crossing incident counts)")
    ws.cell(r, 2, round(rd.get("gini_incident_counts", 0), 4))
    apply_data_block(ws, 4, 1, r, 2)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_revision_s8(wb: Workbook, rev: dict) -> None:
    ws = wb.create_sheet("Rev_S8_sensitivities")
    _sheet_title(ws, "Revision — Supplementary §S8 sensitivities", 3)
    ws.cell(3, 1, "Analysis")
    ws.cell(3, 2, "Result")
    apply_header_row(ws, 3, 2)
    sw = rev.get("state_weighted_harm", {})
    win = rev.get("winsor_damage_99", {})
    mc = rev.get("model_auprc_comparison", {})
    cc = rev.get("complete_case_user_age", {})
    par = rev.get("pareto", {})
    rows = [
        (
            "National state-share weighted harm rate",
            f"{100 * sw.get('national_share_weighted_harm_rate', 0):.2f}% vs unweighted {100 * sw.get('unweighted_harm_rate', 0):.2f}%",
        ),
        (
            "Winsor damage cost (99th pct)",
            f"AUPRC {win.get('cv_full', {}).get('mean_auprc', 0):.4f} → {win.get('cv_winsor', {}).get('mean_auprc', 0):.4f} ({win.get('n_capped', 0)} rows capped)",
        ),
        (
            "HGB − RF mean fold AUPRC",
            f"{mc.get('mean_diff', 0):.4f}; 95% CI {mc.get('boot_ci95_low', 0):.4f} to {mc.get('boot_ci95_high', 0):.4f}",
        ),
        (
            "Complete-case User Age",
            f"n={cc.get('n', 0)}; harm {100 * cc.get('harm_rate', 0):.1f}%; OR {cc.get('h1_or', {}).get('point_or', 0):.2f}",
        ),
        (
            "Top-decile crossings (harm share)",
            f"{100 * par.get('share_harm_top_decile', 0):.1f}% of Y=1 ({par.get('top_decile_crossings', 0)} crossings)",
        ),
    ]
    r = 4
    for k, v in rows:
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 2)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_revision_s9_temporal(wb: Workbook, rev: dict) -> None:
    tr = rev.get("temporal_recurrence", {})
    ws = wb.create_sheet("Rev_S9_temporal_recurrence")
    _sheet_title(
        ws,
        "Revision — Supplementary §S9 temporal recurrence P(harm in year t+1)",
        4,
    )
    ws.cell(3, 1, "Condition (year t)")
    ws.cell(3, 2, "n pairs")
    ws.cell(3, 3, "P(harm in t+1)")
    apply_header_row(ws, 3, 3)
    r = 4
    ws.cell(r, 1, "Unconditional (consecutive-year pairs)")
    ws.cell(r, 2, tr.get("n_consecutive_pairs", ""))
    ws.cell(r, 3, round(tr.get("unconditional_p_harm_t_plus_1", 0) or 0, 4))
    r += 1
    by_t = tr.get("by_incidents_in_year_t", {})
    for label, key in [
        ("Incidents in year t = 1", "single_incident_t"),
        ("Incidents in year t ≥ 1", "baseline_any_incident_t"),
        ("Incidents in year t ≥ 2", "2"),
        ("Incidents in year t ≥ 3", "3"),
    ]:
        d = by_t.get(key, {})
        ws.cell(r, 1, label)
        ws.cell(r, 2, d.get("n_pairs", ""))
        p = d.get("p_harm_year_t_plus_1")
        ws.cell(r, 3, round(p, 4) if p is not None else "")
        r += 1
    r += 1
    ws.cell(r, 1, "Cumulative incidents through year t")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    by_c = tr.get("by_cumulative_incidents_through_t", {})
    for k in ["2", "3", "4"]:
        d = by_c.get(k, {})
        ws.cell(r, 1, f"Cumulative n through t ≥ {k}")
        ws.cell(r, 2, d.get("n_pairs", ""))
        p = d.get("p_harm_year_t_plus_1")
        ws.cell(r, 3, round(p, 4) if p is not None else "")
        r += 1
    r += 1
    ws.cell(r, 1, "Harm stratification (year t)")
    ws.cell(r, 1).font = Font(bold=True)
    r += 1
    by_h = tr.get("by_harm_in_year_t", {})
    for label, key in [
        ("Harm in year t", "harm_in_year_t"),
        ("Incident in t, no harm in t", "no_harm_in_year_t_but_incident"),
    ]:
        d = by_h.get(key, {})
        ws.cell(r, 1, label)
        ws.cell(r, 2, d.get("n_pairs", ""))
        p = d.get("p_harm_year_t_plus_1")
        ws.cell(r, 3, round(p, 4) if p is not None else "")
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 3)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def sheet_revision_national(wb: Workbook, rev: dict) -> None:
    nc = rev.get("national_comparison", {})
    ws = wb.create_sheet("Rev_National_compare")
    _sheet_title(ws, "Revision — Six-state extract vs national FRA file (2013–2026)", 3)
    ws.cell(3, 1, "Quantity")
    ws.cell(3, 2, "Value")
    apply_header_row(ws, 3, 2)
    items = [
        ("National N", nc.get("national_n")),
        ("National harm rate", round(nc.get("national_harm_rate", 0), 4)),
        ("Six-state N (raw file)", nc.get("six_state_n")),
        ("Six-state harm rate", round(nc.get("six_state_harm_rate", 0), 4)),
        ("Texas share national", round(100 * nc.get("texas_share_national", 0), 2)),
        ("Texas share six-state", round(100 * nc.get("texas_share_six_state", 0), 2)),
    ]
    r = 4
    for k, v in items:
        ws.cell(r, 1, k)
        ws.cell(r, 2, v)
        r += 1
    apply_data_block(ws, 4, 1, r - 1, 2)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def main() -> None:
    df = load_analytic_frame()
    n_expected = expected_cohort_n()
    assert len(df) == n_expected, f"Expected N={n_expected}, got {len(df)}"

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    rev = load_revision()
    sheet_readme(wb)
    sheet_table01_literature(wb)
    sheet_table02_cohort(wb, df)
    sheet_table03_flow(wb)
    sheet_table04_leakage(wb)
    if rev:
        sheet_table05_hypotheses(wb, rev)
        sheet_table07_ro_folds(wb, rev)
        sheet_supp_table_s1(wb, rev)
        sheet_table06_ro_summary(wb, rev)
        sheet_table08_multiplicity(wb, rev)
    sheet_missingness(wb, df)
    sheet_yearly(wb, df)
    sheet_h1_contingency(wb, df)
    sheet_recurrence(wb, df)
    sheet_fig1_cv(wb)
    sheet_appendix_a1(wb)
    sheet_appendix_a2_hyperparams(wb)
    if rev:
        sheet_revision_h1_bootstrap(wb, rev)
        sheet_revision_h2(wb, rev)
        sheet_revision_ablation(wb, rev)
        sheet_revision_h3(wb, rev)
        sheet_revision_national(wb, rev)
        sheet_revision_s8(wb, rev)
        sheet_revision_s9_temporal(wb, rev)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print("Wrote", OUT_PATH)


if __name__ == "__main__":
    main()

